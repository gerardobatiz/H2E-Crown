import os, sys, subprocess, shutil, tempfile
from datetime import datetime

BASE    = os.path.dirname(os.path.abspath(__file__))
XLSM    = r"C:\Users\OFFICE DEPOT\OneDrive\Personal GBE\H2E\BCS\Empaque\Formato de Empaque 2E 25-26 (Saladette).xlsm"
DIARIOS = os.path.join(BASE, "Diarios")

def pip_install(pkg):
    subprocess.run([sys.executable, '-m', 'pip', 'install', pkg, '-q'], check=True)

try:
    import openpyxl
except ImportError:
    pip_install('openpyxl')
    import openpyxl

# ── Lectura ───────────────────────────────────────────────────────────────────
def read_sheet():
    if not os.path.exists(XLSM):
        print(f"ERROR: No se encontro el archivo:\n  {XLSM}")
        print("Asegurate de que el archivo principal este en la misma carpeta que este script.")
        sys.exit(1)
    # Copia a temp para evitar PermissionError si el archivo está abierto en Excel / sincronizando OneDrive
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        shutil.copy2(XLSM, tmp_path)
        wb   = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        name = 'p PDF' if 'p PDF' in wb.sheetnames else wb.sheetnames[0]
        ws   = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except:
                pass
    return rows

def g(rows, r, c):
    try:
        v = rows[r][c]
        return v if v is not None else None
    except:
        return None

def cv(v):
    if v is None: return None
    s = str(v).strip()
    return None if s in ('', '-', '—', '#DIV/0!', '#VALUE!', '#N/A', ' -   ', ' - ') else s

def cn(v):
    c = cv(v)
    if c is None: return 0
    try:
        return float(str(c).replace(',', '').replace('$', '').replace('%', '').strip())
    except:
        return 0

def fmt(v, dec=0):
    n = cn(v) if not isinstance(v, (int, float)) else v
    if n == 0 and dec == 0: return '—'
    return f'{n:,.{dec}f}' if dec else f'{int(round(n)):,}'

def fmtusd(v):
    n = cn(v)
    return '—' if n == 0 else f'${n:,.2f}'

def fmtpct(v):
    if v is None: return '—'
    n = cn(v)
    if n == 0: return '—'
    # Excel stores percentages as 0-1 decimals
    if 0 < abs(n) <= 1.0:
        return f'{n * 100:.0f}%'
    return f'{n:.0f}%'

DAYS_ES = {
    'Monday': 'lunes', 'Tuesday': 'martes', 'Wednesday': 'miércoles',
    'Thursday': 'jueves', 'Friday': 'viernes', 'Saturday': 'sábado', 'Sunday': 'domingo',
}
MONTHS_ES = {
    'January': 'enero', 'February': 'febrero', 'March': 'marzo', 'April': 'abril',
    'May': 'mayo', 'June': 'junio', 'July': 'julio', 'August': 'agosto',
    'September': 'septiembre', 'October': 'octubre', 'November': 'noviembre', 'December': 'diciembre',
}

def fmt_date(v):
    if isinstance(v, datetime):
        day_es   = DAYS_ES.get(v.strftime('%A'), v.strftime('%A'))
        month_es = MONTHS_ES.get(v.strftime('%B'), v.strftime('%B'))
        return f'{day_es} {v.day} de {month_es} de {v.year}'
    s = str(v)
    for en, es in {**DAYS_ES, **MONTHS_ES}.items():
        s = s.replace(en, es)
    return s

def esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

# ── Secciones de producto ──────────────────────────────────────────────────────
PRODUCT_ROWS = [
    ('Campo Gusto x 72',    8,  'lb'),
    ('H2E (2da) x 72',     16,  'lb'),
    ('H2E (Orig.) x 80',   24,  'kg'),
    ('Genérica x 80',      32,  'kg'),
    ('1 1/9 x 80',         40,  'kg'),
    ('DF Clam 24 Lb x 50', 48,  'Lb'),
]
SIZES = ['Jb', 'Xl', 'Lg', 'Md', 'Sm', 'Mixto']

# ── Build HTML ────────────────────────────────────────────────────────────────
def build_html(rows):
    # ── Header ────────────────────────────────────────────────────────────────
    dia_num    = cv(g(rows, 1, 9)) or '?'
    rancho     = esc(str(cv(g(rows, 4, 1)) or 'N/A').upper())
    agricultor = esc(cv(g(rows, 2, 1)) or '')
    date_es    = esc(fmt_date(g(rows, 3, 6)))

    # ── Active products ───────────────────────────────────────────────────────
    active = []
    for name, hr, unit in PRODUCT_ROWS:
        if cn(g(rows, hr+7, 2)) > 0 or cn(g(rows, hr+7, 27)) > 0:
            active.append((name, hr, unit))

    # ── Grand totals (row 57) ─────────────────────────────────────────────────
    gt_cajas     = cn(g(rows, 57, 2))
    gt_exp       = cn(g(rows, 58, 9))   # existencia export (row 58)
    gt_nac       = cn(g(rows, 59, 9))   # existencia nacional (row 59)
    gt_exist     = cn(g(rows, 57, 8))
    gt_pals      = cn(g(rows, 57, 9))
    gt_venta     = cn(g(rows, 57, 10))
    gt_rec_caj   = cn(g(rows, 57, 11))
    gt_rec_aprox = cn(g(rows, 57, 12))
    xha_dia      = cn(g(rows, 59, 2))

    # ── Recepción ─────────────────────────────────────────────────────────────
    ant_bins = cn(g(rows, 63, 1));  ant_kg   = cn(g(rows, 63, 2))
    bins_cnt = cn(g(rows, 64, 1));  bins_kg  = cn(g(rows, 64, 2))
    cajas_rx = cn(g(rows, 65, 1));  cajas_kg = cn(g(rows, 65, 2))
    total_kg = cn(g(rows, 66, 2))
    # ── Peso Empacado ─────────────────────────────────────────────────────────
    pct_1ra  = g(rows, 67, 2)
    pct_2da  = g(rows, 68, 2)
    kg_emp     = cn(g(rows, 62, 6))
    pct_emp_kg = g(rows, 62, 7)
    pct_tamano = g(rows, 63, 6)
    pct_jbxl   = g(rows, 63, 7)
    cxcubeta   = cn(g(rows, 64, 6))
    cxbin_calc = gt_cajas / bins_cnt if bins_cnt else 0

    # ── Por pasar y reproc ────────────────────────────────────────────────────
    ppr_bins  = cn(g(rows, 67, 6));  ppr_bins_kg  = cn(g(rows, 67, 7))
    ppr_cajas = cn(g(rows, 68, 6));  ppr_cajas_kg = cn(g(rows, 68, 7))

    # ── Recuperación (from first active product, first row with precio) ───────
    precio_caja = rec_caja = recuperacion = 0
    if active:
        _, hr, _ = active[0]
        for i in range(6):
            p = cn(g(rows, hr+1+i, 10))
            if p > 0:
                precio_caja = p
                rec_caja    = cn(g(rows, hr+1+i, 11))
                break
        recuperacion = gt_rec_aprox

    # ── Semana ────────────────────────────────────────────────────────────────
    cajas_sem = cn(g(rows, 71, 1))
    xha_sem   = cn(g(rows, 71, 2))
    jbxl_sem  = g(rows, 72, 1)

    # ── Acum x ha season totals ───────────────────────────────────────────────
    acum_xha_current = 0
    acum_season_rows = []
    for ri in range(75, 82):
        name_v = cv(g(rows, ri, 0))
        val_v  = g(rows, ri, 1)
        vs_v   = g(rows, ri, 2)
        if not name_v: continue
        n = cn(val_v)
        if n == 0: continue
        vs_s = fmtpct(vs_v) if cv(vs_v) else '—'
        if ri == 75: acum_xha_current = n
        acum_season_rows.append((esc(name_v), fmt(n), vs_s))

    # ── Product sections HTML ─────────────────────────────────────────────────
    sections_html = ''
    acum_html     = ''

    for prod_name, hr, unit in active:
        trows = ''
        for i, sz in enumerate(SIZES):
            sr    = hr + 1 + i
            cajas = cn(g(rows, sr, 2))
            exist = cn(g(rows, sr, 8))
            nac_d = cn(g(rows, sr, 7))
            exp_d = cn(g(rows, sr, 6))
            if cajas == 0 and exist == 0 and nac_d == 0: continue
            peso  = cv(g(rows, sr, 1)) or ''
            pct   = fmtpct(g(rows, sr, 3))
            bajas = cn(g(rows, sr, 5))
            pals  = cn(g(rows, sr, 9))
            trows += f'''<tr>
              <td>{sz}</td>
              <td class="r">{peso}</td>
              <td class="r b">{fmt(cajas) if cajas else "—"}</td>
              <td class="r">{pct}</td>
              <td class="r">{fmt(bajas) if bajas else "—"}</td>
              <td class="r">{fmt(exp_d) if exp_d else "—"}</td>
              <td class="r">{fmt(nac_d) if nac_d else "—"}</td>
              <td class="r">{fmt(exist) if exist else "—"}</td>
              <td class="r">{f"{pals:.1f}" if pals else "—"}</td>
            </tr>'''

        tr = hr + 7
        trows += f'''<tr class="tot">
          <td><b>TOTALES</b></td>
          <td class="r">—</td>
          <td class="r b">{fmt(cn(g(rows, tr, 2)))}</td>
          <td class="r">100%</td>
          <td class="r">{fmt(cn(g(rows, tr, 5))) if cn(g(rows, tr, 5)) else "—"}</td>
          <td class="r">{fmt(cn(g(rows, tr, 6))) if cn(g(rows, tr, 6)) else "—"}</td>
          <td class="r">{fmt(cn(g(rows, tr, 7))) if cn(g(rows, tr, 7)) else "—"}</td>
          <td class="r">{fmt(cn(g(rows, tr, 8)))}</td>
          <td class="r">{f"{cn(g(rows, tr, 9)):.1f}" if cn(g(rows, tr, 9)) else "—"}</td>
        </tr>'''

        sections_html += f'''
        <p class="prod-lbl">{esc(prod_name)} · <em>{unit}</em></p>
        <table class="mt">
          <thead>
            <tr>
              <th colspan="4" class="se">EMPAQUE</th>
              <th class="sb">&nbsp;</th>
              <th colspan="2" class="sb">EMBARQUE</th>
              <th colspan="2" class="sx">EXISTENCIA</th>
            </tr>
            <tr class="sh">
              <th>Descripción</th><th class="r">Peso {unit}</th>
              <th class="r">Cajas Emp</th><th class="r">%</th>
              <th class="r">Bajas / Reemp</th>
              <th class="r">Exp</th><th class="r">Nac</th>
              <th class="r">Existencia</th><th class="r">Pallets</th>
            </tr>
          </thead>
          <tbody>{trows}</tbody>
        </table>'''

        # Acum table for this product
        arows = ''
        for i, sz in enumerate(SIZES):
            sr  = hr + 1 + i
            ca  = cn(g(rows, sr, 27))
            if ca == 0: continue
            ea  = cn(g(rows, sr, 33))
            ex  = cn(g(rows, sr, 8))
            na  = ca - ea - ex
            pct = fmtpct(g(rows, sr, 28))
            arows += f'''<tr>
              <td>{sz}</td>
              <td class="r b">{fmt(ca)}</td>
              <td class="r">{pct}</td>
              <td class="r">{"—" if ea == 0 else fmt(ea)}</td>
              <td class="r">{"—" if na <= 0 else fmt(na)}</td>
              <td class="r">{fmt(ex)}</td>
            </tr>'''

        tr   = hr + 7
        ca_t = cn(g(rows, tr, 27))
        ea_t = cn(g(rows, tr, 33))
        ex_t = cn(g(rows, tr, 8))
        na_t = ca_t - ea_t - ex_t
        arows += f'''<tr class="tot">
          <td><b>TOTALES</b></td>
          <td class="r b">{fmt(ca_t)}</td>
          <td class="r">100%</td>
          <td class="r">{"—" if ea_t == 0 else fmt(ea_t)}</td>
          <td class="r">{"—" if na_t <= 0 else fmt(na_t)}</td>
          <td class="r">{fmt(ex_t)}</td>
        </tr>'''

        acum_html += f'''
        <div class="acum-wrap">
          <div class="acum-ttl">Acumulados Temporada — {esc(prod_name)}</div>
          <table class="at">
            <thead><tr>
              <th style="text-align:left">Tamaño</th>
              <th class="r">Cajas Acum</th>
              <th class="r">% del Total</th>
              <th class="r">Exp Acum</th>
              <th class="r">Nac Acum</th>
              <th class="r">Existencia</th>
            </tr></thead>
            <tbody>{arows}</tbody>
          </table>
        </div>'''

    # ── Semana row con acum xha ────────────────────────────────────────────────
    season_lbl = esc(cv(g(rows, 75, 0)) or '2E 25-26')

    # ── Comparación x ha vs temporadas ───────────────────────────────────────
    vs_rows_html = ''
    for i, (sname, sval, svs) in enumerate(acum_season_rows):
        cls = ' class="vs-current"' if i == 0 else ''
        neg = ' class="r neg"' if svs.startswith('-') else ' class="r pos"' if svs not in ('—', '') else ' class="r"'
        vs_rows_html += f'<tr{cls}><td>{sname}</td><td class="r b">{sval}</td><td{neg}>{svs}</td></tr>'

    gen = datetime.now().strftime('%d/%m/%Y %H:%M')

    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><style>
*{{box-sizing:border-box;margin:0;padding:0;font-family:Arial,sans-serif;font-size:9.5px}}
body{{padding:10px 14px;background:#fff}}
.hdr{{background:#1b4332;color:#fff;padding:8px 12px;border-radius:4px;
      display:flex;justify-content:space-between;align-items:center;margin-bottom:3px}}
.hl h1{{font-size:15px;font-weight:bold}}
.hl p{{font-size:9px;opacity:.75;margin-top:2px}}
.hr{{text-align:right}}
.hr .dl{{font-size:9px;opacity:.7}}
.hr .dn{{font-size:34px;font-weight:bold;color:#ffd166;line-height:1}}
.hr .rn{{font-size:12px;font-weight:bold;margin-top:1px}}
.date{{text-align:center;font-size:11px;color:#333;margin:5px 0;font-style:italic}}
.prod-lbl{{font-size:9px;color:#666;font-style:italic;margin:6px 0 2px}}
.mt{{width:100%;border-collapse:collapse;margin-bottom:2px}}
.mt th{{padding:2px 4px;text-align:center;color:#fff;font-size:8.5px}}
.mt .se{{background:#5c3015}}.mt .sb{{background:#7b4820}}.mt .sx{{background:#2d6a4f}}
.mt .sh th{{background:#444;font-size:8.5px;padding:2px 4px}}
.mt td{{padding:2px 4px;border-bottom:1px solid #eee}}
.mt tr:nth-child(even) td{{background:#f8f8f8}}
.mt .tot td{{background:#e4e4e4;font-weight:bold;border-top:1px solid #bbb}}
.r{{text-align:right}}.b{{font-weight:bold}}
.cxb{{font-size:8.5px;color:#555;display:flex;justify-content:space-between;
      margin:2px 0 8px;padding:3px 6px;background:#f5f5f5;border-radius:3px}}
.mid{{display:flex;gap:6px;margin:7px 0}}
.mbox{{flex:1;border:1px solid #ddd;border-radius:4px;overflow:hidden}}
.mbox-t{{background:#2d6a4f;color:#fff;padding:3px 7px;font-size:8.5px;
         font-weight:bold;text-align:center}}
.mbox table{{width:100%;border-collapse:collapse}}
.mbox td{{padding:2px 7px;border-bottom:1px solid #f0f0f0;font-size:9px}}
.mbox .rv{{text-align:right;font-weight:bold}}
.mbox .sub{{background:#f0f7f4;font-size:8px;color:#555}}
.mbox .sub td{{padding:1px 7px}}
.hl2{{color:#1b4332;font-size:13px;font-weight:bold}}
.semrow{{display:flex;gap:6px;margin:7px 0}}
.sembox{{background:#1b4332;color:#fff;border-radius:4px;padding:7px 10px;flex:1}}
.sembox .st{{font-size:8px;text-align:center;opacity:.75;font-weight:bold;
             text-transform:uppercase;letter-spacing:.05em;margin-bottom:5px}}
.sr{{display:flex;justify-content:space-between;font-size:9px;padding:2px 0}}
.sr .sv{{font-size:13px;font-weight:bold;color:#95d5b2}}
.acumbox{{background:#2d6a4f;color:#fff;border-radius:4px;padding:7px 10px;
          flex:1;text-align:center}}
.acumbox .at0{{font-size:8px;opacity:.75;text-transform:uppercase;
               font-weight:bold;letter-spacing:.05em}}
.acumbox .an{{font-size:44px;font-weight:bold;color:#b7e4c7;line-height:1.1}}
.acumbox .as{{font-size:9px;opacity:.75;margin-bottom:3px}}
.acum-wrap{{margin:7px 0}}
.acum-ttl{{background:#1b4332;color:#fff;padding:3px 8px;font-size:9px;
           font-weight:bold;border-radius:4px 4px 0 0;text-align:center}}
.at{{width:100%;border-collapse:collapse}}
.at th{{background:#2d6a4f;color:#fff;padding:2px 6px;font-size:8.5px}}
.at td{{padding:2px 6px;border-bottom:1px solid #eee;font-size:9px}}
.at tr:nth-child(even) td{{background:#f8f8f8}}
.at .tot td{{background:#e4e4e4;font-weight:bold}}
.vs-wrap{{margin:7px 0}}
.vs-ttl{{background:#1b4332;color:#fff;padding:3px 8px;font-size:9px;
         font-weight:bold;border-radius:4px 4px 0 0;text-align:center}}
.vs-tbl{{width:100%;border-collapse:collapse}}
.vs-tbl th{{background:#2d6a4f;color:#fff;padding:2px 8px;font-size:8.5px}}
.vs-tbl th:last-child{{text-align:right}}
.vs-tbl td{{padding:2px 8px;border-bottom:1px solid #eee;font-size:9px}}
.vs-tbl .vs-current td{{background:#d8f3dc;font-weight:bold}}
.vs-tbl .neg{{color:#c0392b}}
.vs-tbl .pos{{color:#1b4332}}
.foot{{text-align:center;font-size:8px;color:#aaa;margin-top:10px;
       border-top:1px solid #eee;padding-top:4px}}
</style></head><body>

<div class="hdr">
  <div class="hl">
    <h1>Reporte Diario de Empaque — Saladette</h1>
    <p>H2E México SA de CV &nbsp;·&nbsp; Agricultor: {agricultor}</p>
  </div>
  <div class="hr">
    <div class="dl">Dia Empaque #</div>
    <div class="dn">{dia_num}</div>
    <div class="rn">Rancho: <b>{rancho}</b></div>
  </div>
</div>
<div class="date">{date_es}</div>

{sections_html}

<div class="cxb">
  <span>Cajas × bin: <b>{cxbin_calc:.1f}</b> &nbsp;·&nbsp; × ha <b>{fmt(xha_dia) if xha_dia else "—"}</b> &nbsp;·&nbsp; 1ra <b>{fmtpct(pct_1ra)}</b> &nbsp;·&nbsp; 2da/Nac <b>{fmtpct(pct_2da)}</b></span>
  <span>Existencia — Exp <b>{fmt(gt_exp) if gt_exp else "0"}</b> · Nac <b>{fmt(gt_nac) if gt_nac else "0"}</b></span>
</div>

<div class="mid">
  <div class="mbox">
    <div class="mbox-t">Recepción</div>
    <table>
      <tr class="sub"><td colspan="2">kg (Aprox)</td></tr>
      <tr><td>Ant y Reemp</td><td class="rv">{fmt(ant_bins, 1) if ant_bins else "0"} &nbsp; {fmt(ant_kg) if ant_kg else "—"}</td></tr>
      <tr><td>Bins</td><td class="rv">{fmt(bins_cnt, 1) if bins_cnt else "0"} &nbsp; {fmt(bins_kg) if bins_kg else "—"}</td></tr>
      <tr><td>Cajas</td><td class="rv">{fmt(cajas_rx) if cajas_rx else "0"} &nbsp; {fmt(cajas_kg) if cajas_kg else "—"}</td></tr>
      <tr><td><b>Total kg</b></td><td class="rv"><b>{fmt(total_kg)}</b></td></tr>
    </table>
  </div>

  <div class="mbox">
    <div class="mbox-t">Peso Empacado</div>
    <table>
      <tr><td>kg Recibidos</td><td class="rv">{fmt(total_kg)}</td></tr>
      <tr><td>kg Empacados</td><td class="rv">{fmt(kg_emp)}</td></tr>
      <tr><td>% emp kg</td><td class="rv">{fmtpct(pct_emp_kg)}</td></tr>
      <tr><td>% Tamaño Grandes</td><td class="rv">{fmtpct(pct_tamano)}</td></tr>
      <tr><td>% Jb / Xl</td><td class="rv">{fmtpct(pct_jbxl)}</td></tr>
      <tr><td>Cajas × bin</td><td class="rv">{cxbin_calc:.1f}</td></tr>
      <tr><td>Cajas × cubeta</td><td class="rv">{cxcubeta:.2f}</td></tr>
    </table>
    <div class="mbox-t" style="margin-top:4px;border-radius:0">Por pasar y reproc</div>
    <table>
      <tr><td>Bins</td><td class="rv">{fmt(ppr_bins) if ppr_bins else "0"} &nbsp; {fmt(ppr_bins_kg) if ppr_bins_kg else "—"}</td></tr>
      <tr><td>Cajas</td><td class="rv">{fmt(ppr_cajas) if ppr_cajas else "0"} &nbsp; {fmt(ppr_cajas_kg) if ppr_cajas_kg else "—"}</td></tr>
    </table>
  </div>

  <div class="mbox">
    <div class="mbox-t">Recuperación Est.</div>
    <table>
      <tr><td><b>Venta estimada</b></td><td class="rv hl2"><b>{fmtusd(gt_venta)}</b></td></tr>
      <tr><td>Precio est. / caja</td><td class="rv hl2">{fmtusd(precio_caja)}</td></tr>
      <tr><td>Rec × caja</td><td class="rv hl2">{fmtusd(rec_caja)}</td></tr>
      <tr><td><b>Recuperación est.</b></td><td class="rv hl2"><b>{fmtusd(recuperacion)}</b></td></tr>
    </table>
  </div>
</div>

<div class="semrow">
  <div class="sembox">
    <div class="st">Semana</div>
    <div class="sr"><span>Cajas</span><span class="sv">{fmt(cajas_sem)}</span></div>
    <div class="sr"><span>× ha</span><span class="sv">{fmt(xha_sem) if xha_sem else "—"}</span></div>
    <div class="sr"><span>% Jb / Xl</span><span class="sv">{fmtpct(jbxl_sem)}</span></div>
  </div>
  <div class="acumbox">
    <div class="at0">Acum. Temporada {season_lbl}</div>
    <div class="as">Acum. a fecha × ha</div>
    <div class="an">{fmt(acum_xha_current)}</div>
  </div>
</div>

{acum_html}

<div class="vs-wrap">
  <div class="vs-ttl">Acum. a fecha × ha — comparación vs temporadas anteriores</div>
  <table class="vs-tbl">
    <thead><tr><th style="text-align:left">Temporada</th><th style="text-align:right">Acum × ha</th><th style="text-align:right">vs actual</th></tr></thead>
    <tbody>{vs_rows_html}</tbody>
  </table>
</div>

<div class="foot">H2E México SA de CV &nbsp;·&nbsp; Generado: {gen}</div>
</body></html>'''


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    rows    = read_sheet()
    html    = build_html(rows)
    dia_num    = cv(g(rows, 1, 9)) or '0'
    fecha_emp  = g(rows, 3, 6)
    fecha_str  = fecha_emp.strftime('%d-%m-%y') if isinstance(fecha_emp, datetime) else datetime.now().strftime('%d-%m-%y')
    outfile = os.path.join(DIARIOS, f"Reporte Saladette Dia {dia_num} {fecha_str}.pdf")
    os.makedirs(DIARIOS, exist_ok=True)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        pip_install('playwright')
        subprocess.run([sys.executable, '-m', 'playwright', 'install', 'chromium', '--with-deps'], check=True)
        from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page    = browser.new_page()
        page.set_content(html, wait_until='networkidle')
        page.pdf(path=outfile, format='Letter', print_background=True,
                 margin={'top': '10mm', 'bottom': '10mm', 'left': '10mm', 'right': '10mm'})
        browser.close()

    print(f"Reporte generado: {os.path.basename(outfile)}")
    return outfile


def make_test_rows():
    """Filas sintéticas con datos del Día 9 (11-05-26) para prueba."""
    R = 101
    C = 40
    rows = [[None]*C for _ in range(R)]
    def s(r, c, v): rows[r][c] = v

    s(0, 0, 'H2E México SA de CV')
    s(1, 9, 9)
    s(2, 1, 'H2E México SA de CV')
    s(3, 6, datetime(2026, 5, 11))
    s(4, 1, 'Mayela')

    # H2E (Orig.) x 80 at hr=24
    hr = 24
    s(hr, 0, 'H2E (Orig.) x 80'); s(hr, 1, 'kg')
    data = [
        # sz,  peso,  cajas, pct,   bajas, exp, nac, exist, pallets, precio, rec_caj, rec_aprox, ca,    pct_a, ea,  na
        ('Jb', 12.5,  206,   0.72,  0, 0, 0, 434,  6.0,  430, 394, 81164,  1234, 0.821, 0, 800),
        ('Xl', 12.5,   42,   0.15,  0, 0, 0, 101,  1.4,  430, 394, 16548,   141, 0.094, 0,  40),
        ('Lg', 12.5,   30,   0.10,  0, 0, 0,  48,  0.7,  430, 394, 11820,    75, 0.050, 0,  27),
        ('Md', 12.5,    7,   0.02,  0, 0, 0,  17,  0.2,  430, 394,  2758,    33, 0.022, 0,  16),
        ('Sm', 12.5,    2,   0.01,  0, 0, 0,  11,  0.2,  430, 394,   788,    19, 0.013, 0,   8),
        ('Mixto',12.5,  0,   0,     0, 0, 0,   0,  0.0,    0,   0,     0,     0, 0,     0,   0),
    ]
    for i, (sz, peso, cajas, pct, bajas, exp, nac, exist, pals, precio, rccaj, rcapr, ca, pca, ea, na) in enumerate(data):
        sr = hr + 1 + i
        s(sr, 0, sz); s(sr, 1, peso); s(sr, 2, cajas); s(sr, 3, pct)
        s(sr, 5, bajas); s(sr, 6, exp); s(sr, 7, nac); s(sr, 8, exist)
        s(sr, 9, pals); s(sr, 10, precio); s(sr, 11, rccaj); s(sr, 12, rcapr)
        s(sr, 27, ca); s(sr, 28, pca); s(sr, 33, ea)

    # Totals row for product (hr+7 = row 31)
    tr = hr + 7
    s(tr, 2, 287); s(tr, 5, 0); s(tr, 6, 0); s(tr, 7, 0)
    s(tr, 8, 611); s(tr, 9, 8.5); s(tr, 10, 123410); s(tr, 11, 394); s(tr, 12, 113078)
    s(tr, 27, 1502); s(tr, 28, 1.0); s(tr, 33, 0)

    # Grand totals row 57
    s(57, 2, 287); s(57, 6, 0); s(57, 7, 0); s(57, 8, 611)
    s(57, 9, 8.5); s(57, 10, 123410); s(57, 11, 394); s(57, 12, 113078)

    # row 58: exp existencia, row 59: x ha / nac existencia
    s(58, 8, 'Exp'); s(58, 9, 0)
    s(59, 1, 'x ha'); s(59, 2, 230); s(59, 8, 'Nac'); s(59, 9, 611)

    # Recepción
    s(63, 1, 0);    s(63, 2, 0)       # Ant y Reemp
    s(64, 1, 13.3); s(64, 2, 5114)    # Bins
    s(65, 1, 0);    s(65, 2, 0)       # Cajas
    s(66, 2, 5114)                     # Total kg

    s(67, 2, 0)     # 1ra  (0 = 0%)
    s(68, 2, 1)     # 2da/Nac (1 = 100%)

    # Peso empacado
    s(62, 6, 3588); s(62, 7, 0.70)   # kg emp, % emp kg
    s(63, 6, 0.97); s(63, 7, 0.86)   # % tamaño, % Jb/Xl
    s(64, 6, 0.72)                    # cajas × cubeta
    s(65, 6, 22)                      # cajas × bin

    # Por pasar y reproc
    s(67, 6, 0); s(67, 7, 0)
    s(68, 6, 0); s(68, 7, 0)

    # Semana
    s(71, 1, 287); s(71, 2, 230)
    s(72, 1, 0.86)

    # Acum a fecha x ha
    s(74, 0, 'Acum a fecha x ha'); s(74, 2, 'vs')
    s(75, 0, 'Actual');   s(75, 1, 1202)
    s(76, 0, '2E 2025');  s(76, 1, 602);  s(76, 2, 1.0)
    s(77, 0, '2E 2024');  s(77, 1, 2016); s(77, 2, -0.404)
    s(78, 0, '2E 2023');  s(78, 1, 2476); s(78, 2, -0.514)

    return rows


if __name__ == '__main__':
    import sys
    try:
        if '--test' in sys.argv:
            rows   = make_test_rows()
            html   = build_html(rows)
            today  = datetime.now().strftime('%d-%m-%y')
            outfile = os.path.join(DIARIOS, f"Reporte Saladette Dia 9 {today} (Prueba).pdf")
            os.makedirs(DIARIOS, exist_ok=True)
            try:
                from playwright.sync_api import sync_playwright
            except ImportError:
                pip_install('playwright')
                subprocess.run([sys.executable, '-m', 'playwright', 'install', 'chromium', '--with-deps'], check=True)
                from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page    = browser.new_page()
                page.set_content(html, wait_until='networkidle')
                page.pdf(path=outfile, format='Letter', print_background=True,
                         margin={'top': '10mm', 'bottom': '10mm', 'left': '10mm', 'right': '10mm'})
                browser.close()
            print(f"Prueba generada: {os.path.basename(outfile)}")
        else:
            main()
    except Exception as e:
        print(f"\nERROR: {e}")
    input("\nPresiona Enter para cerrar...")
